from __future__ import annotations
from typing import cast

# build_gw2_trade_ledger.py

import os
import math
from typing import Dict, List, Tuple, Any, Iterable
from datetime import datetime, timezone

import pandas as pd

# Import your existing client
from gw2 import GW2  # assumes gw2.py is in the same folder or on PYTHONPATH


# -------------------------- helpers --------------------------

COPPER_PER_GOLD = 10000


def copper_to_gold_float(copper: int) -> float:
    return round(copper / COPPER_PER_GOLD, 6)


def parse_iso8601(s: str) -> datetime:
    # GW2 API returns ISO8601 UTC like "2024-05-11T17:43:01Z"
    return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)


def chunked(it: Iterable[int], size: int = 200):
    buf: List[int] = []
    for x in it:
        buf.append(int(x))
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


# --------------------- item name resolution ---------------------


def fetch_item_names(gw2: GW2, ids: Iterable[int]) -> Dict[int, str]:
    """
    Bulk-fetch item names via /v2/items?ids=...
    Returns {item_id: name}
    """
    out: Dict[int, str] = {}
    for chunk in chunked(ids, 200):
        params = {"ids": ",".join(map(str, chunk))}
        data = gw2._request("GET", "/items", params=params)  # reuse client's session
        for it in data:
            item_id = int(it.get("id"))
            name = it.get("name") or f"Item {item_id}"
            out[item_id] = name
    return out


# ----------------------- data acquisition -----------------------


def fetch_all_transactions(
    gw2: GW2, scope: str, side: str, page_size: int = 200
) -> List[Dict[str, Any]]:
    """
    Pulls all pages for /v2/commerce/transactions/{scope}/{side}
    """
    page = 0
    out: List[Dict[str, Any]] = []
    while True:
        batch = gw2.transactions(scope, side, page=page, page_size=page_size)
        if not batch:
            break
        out.extend(batch)
        if len(batch) < page_size:
            break
        page += 1
    return out


from collections import defaultdict


def coalesce_buys_by_day(buys, tz: str = "US/Eastern"):
    """
    Coalesce partial buy fills by (item_id, unit price, calendar day in `tz`).
    - Sum quantities.
    - Use quantity-weighted average timestamp for 'purchased'.
    Returns a new list of dicts compatible with the FIFO matcher.
    """
    if not buys:
        return buys

    def _as_utc_dt(s: str) -> datetime:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)

    def _iso_z(dt: datetime) -> str:
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # Build groups keyed by (item_id, price, YYYY-MM-DD in tz)
    groups = defaultdict(lambda: {"qty": 0, "wts": 0.0, "any_id": 0})
    for b in buys:
        item = int(b["item_id"])
        price = int(b["price"])
        qty = int(b["quantity"])
        dt_utc = _as_utc_dt(b["purchased"])
        # Convert to chosen tz and take the calendar date
        local = pd.Timestamp(dt_utc).tz_convert(tz)
        ymd = str(local.date())

        key = (item, price, ymd)
        groups[key]["qty"] = int(groups[key]["qty"]) + int(qty)
        ts = float(dt_utc.timestamp())
        groups[key]["wts"] = float(groups[key]["wts"]) + float(qty) * ts
        if not groups[key]["any_id"]:
            groups[key]["any_id"] = int(b.get("id", 0))

    # Emit merged rows: timestamp = qty-weighted average of the day's fills (in UTC)
    merged = []
    for (item, price, _ymd), agg in groups.items():
        qty = agg["qty"]
        avg_ts = float(agg["wts"]) / float(max(1, int(qty)))
        avg_dt_utc = datetime.fromtimestamp(avg_ts, tz=timezone.utc)
        merged.append(
            {
                "id": agg["any_id"],
                "item_id": item,
                "price": price,
                "quantity": qty,
                "purchased": _iso_z(avg_dt_utc),
            }
        )

    # Keep deterministic order
    merged.sort(key=lambda x: (x["item_id"], x["price"], x["purchased"]))
    print(f"Coalesced buys into {len(merged)} daily lots (from {len(buys)} fills).")
    return merged


def coalesce_sells_by_day(sells, tz: str = "US/Eastern"):
    """
    Merge sell fills by (item_id, unit price, calendar day in `tz`).
    Produces aggregated rows with summed quantity and qty-weighted average timestamp.
    Creates a stable synthetic id 'item-price-YYYYMMDD' for de-dup and ledger rows.
    """
    if not sells:
        return sells

    def _as_utc_dt(s: str) -> datetime:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)

    groups: "defaultdict[tuple[int, int, str], dict[str, Any]]" = defaultdict(
        lambda: {"qty": 0, "wts": 0.0, "ids": []}
    )
    for s in sells:
        item = int(s["item_id"])
        price = int(s["price"])
        qty = int(s["quantity"])
        dt_utc = _as_utc_dt(s["purchased"])
        local = pd.Timestamp(dt_utc).tz_convert(tz)
        ymd = str(local.date())  # YYYY-MM-DD
        key = (item, price, ymd)
        groups[key]["qty"] = int(groups[key]["qty"]) + int(qty)
        ts = float(dt_utc.timestamp())
        groups[key]["wts"] = float(groups[key]["wts"]) + float(qty) * ts
        if "id" in s:
            groups[key]["ids"].append(str(s["id"]))

    merged = []
    for (item, price, ymd), agg in groups.items():
        qty = int(agg.get("qty", 0))
        wts = float(agg.get("wts", 0.0))
        avg_ts = wts / float(max(1, qty))
        avg_dt_utc = datetime.fromtimestamp(
            float(avg_ts), tz=timezone.utc
        )  # pyright: ignore[reportArgumentType]
        synthetic_id = f"{item}-{price}-{ymd.replace('-', '')}"
        merged.append(
            {
                "id": synthetic_id,  # string id for dedup
                "item_id": item,
                "price": price,
                "quantity": qty,
                "purchased": avg_dt_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "sell_ids": ",".join(agg["ids"]) if agg["ids"] else None,
            }
        )

    merged.sort(key=lambda x: (x["item_id"], x["price"], x["purchased"]))
    print(f"Coalesced sells into {len(merged)} daily lots (from {len(sells)} fills).")
    return merged


# ----------------------- matching engine -----------------------


def fifo_match_and_rows(
    buys: List[Dict[str, Any]],
    sells: List[Dict[str, Any]],
    name_lookup: Dict[int, str],
) -> List[Dict[str, Any]]:
    """
    For each sell, consume buy lots in FIFO order for the same item_id.
    Build a row per sell with:
      Sell Date (UTC), Item, Quantity, Avg Buy Price (g), Sell Price (g),
      Proceeds (g), Cost (g), Profit (g), Hold Days, sell_id, item_id

    Net proceeds factor = 0.85 (5% listing + 10% exchange).
    Hold Days uses a qty-weighted average buy timestamp of the matched lots.
    """
    # Build FIFO buy lots per item
    buys_by_item: Dict[int, List[List[Any]]] = {}
    for b in sorted(
        buys,
        key=lambda x: (int(x["item_id"]), parse_iso8601(x["purchased"]).timestamp()),
    ):
        item_id = int(b["item_id"])
        lot = [
            int(b["quantity"]),  # remaining qty
            int(b["price"]),  # copper per unit
            parse_iso8601(b["purchased"]),  # buy time
            int(b["id"]),  # buy id
        ]
        buys_by_item.setdefault(item_id, []).append(lot)

    # Sort sells by time asc for proper FIFO consumption
    sells_sorted = sorted(
        sells, key=lambda x: parse_iso8601(x["purchased"]).timestamp()
    )

    rows: List[Dict[str, Any]] = []
    fee_factor = 0.85

    for s in sells_sorted:
        item_id = int(s["item_id"])
        qty_to_match = int(s["quantity"])
        sell_unit_copper = int(s["price"])
        sell_time = parse_iso8601(s["purchased"])
        sell_id = str(s["id"])
        item_name = name_lookup.get(item_id, f"Item {item_id}")

        lots = buys_by_item.get(item_id, [])

        # For weighted-average buy time & average buy price
        total_cost_copper = 0
        matched_qty = 0
        weighted_buy_time_seconds = 0.0
        had_any_match = False

        while qty_to_match > 0 and lots:
            lot = lots[0]
            lot_qty, lot_price_copper, lot_time = lot[0], lot[1], lot[2]
            take = min(qty_to_match, lot_qty)
            had_any_match = True

            total_cost_copper += take * lot_price_copper
            matched_qty += take
            weighted_buy_time_seconds += take * lot_time.timestamp()

            lot[0] -= take
            qty_to_match -= take
            if lot[0] == 0:
                lots.pop(0)

        if matched_qty == 0:
            # Skip this sell completely — no parallel buy order (not a flip)
            continue

        avg_buy_per_unit_copper = math.floor(total_cost_copper / max(1, matched_qty))

        # Hold days (qty-weighted avg buy time) — if nothing matched, set NaN
        if matched_qty > 0 and weighted_buy_time_seconds > 0:
            avg_buy_ts = weighted_buy_time_seconds / matched_qty
            avg_buy_dt = datetime.fromtimestamp(avg_buy_ts, tz=timezone.utc)
            hold_days = (sell_time - avg_buy_dt).total_seconds() / 86400.0
        else:
            avg_buy_dt = None
            hold_days = float("nan")

        # Compute proceeds & P/L using the sell's full quantity
        qty = int(s["quantity"])
        gross_copper = qty * sell_unit_copper
        proceeds_copper = int(round(gross_copper * fee_factor))
        cost_copper = qty * avg_buy_per_unit_copper
        profit_copper = proceeds_copper - cost_copper

        rows.append(
            {
                "Sell Date (UTC)": sell_time,
                "Item": item_name,
                "Quantity": qty,
                "Avg Buy Price (g)": copper_to_gold_float(avg_buy_per_unit_copper),
                "Sell Price (g)": copper_to_gold_float(sell_unit_copper),
                "Proceeds (g)": copper_to_gold_float(proceeds_copper),
                "Cost (g)": copper_to_gold_float(cost_copper),
                "Profit (g)": copper_to_gold_float(profit_copper),
                "ROI": (
                    round(
                        (
                            copper_to_gold_float(profit_copper)
                            / max(1e-12, copper_to_gold_float(cost_copper))
                        )
                        * 100,
                        6,
                    )
                    if cost_copper > 0
                    else None
                ),
                "Hold Days": round(hold_days, 3) if not math.isnan(hold_days) else None,
                "sell_id": sell_id,
                "item_id": item_id,
            }
        )

    return rows


# ----------------------- Excel persistence -----------------------


def append_to_excel(ledger_rows: List[Dict[str, Any]], xlsx_path: str) -> None:
    """
    Append-only semantics using the unique sell_id:
      - If file exists, load it and drop any new rows whose sell_id is already present.
      - Combine, recompute running Total Profit on the combined set (chronological),
        then save back to the same file (logical append; no duplicates).
    """
    new_df = pd.DataFrame(ledger_rows)

    if new_df.empty:
        print("No sell transactions found. Nothing to append.")
        return

    # Convert to US/Eastern, then drop tz to make it naive; rename column to EST
    new_df["Sell Date (UTC)"] = (
        pd.to_datetime(new_df["Sell Date (UTC)"], utc=True)
        .dt.tz_convert("US/Eastern")
        .dt.tz_localize(None)
    )
    new_df = new_df.rename(columns={"Sell Date (UTC)": "Sell Date (EST)"})
    new_df["sell_id"] = new_df["sell_id"].astype(str)

    if os.path.exists(xlsx_path):
        try:
            old_df = pd.read_excel(xlsx_path, engine="openpyxl")
        except Exception:
            old_df = pd.read_excel(xlsx_path)
        # Normalize legacy files: handle either UTC or EST naming
        if "Sell Date (EST)" in old_df.columns:
            old_df["Sell Date (EST)"] = pd.to_datetime(
                old_df["Sell Date (EST)"], errors="coerce"
            )
        elif "Sell Date (UTC)" in old_df.columns:
            old_df["Sell Date (UTC)"] = (
                pd.to_datetime(old_df["Sell Date (UTC)"], utc=True, errors="coerce")
                .dt.tz_convert("US/Eastern")
                .dt.tz_localize(None)
            )
            old_df = old_df.rename(columns={"Sell Date (UTC)": "Sell Date (EST)"})

        # ensure sell_id exists even on old files created before column existed
        if "sell_id" not in old_df.columns:
            old_df["sell_id"] = pd.NA

        existing_ids = set(old_df["sell_id"].dropna().astype(str).tolist())
        new_df = new_df[~new_df["sell_id"].isin(existing_ids)]
        combined = pd.concat([old_df, new_df], ignore_index=True, copy=False)
    else:
        combined = new_df

    # How many NEW rows are being added in this run (after de-dup)
    added_rows = len(new_df)
    print(f"New rows added this run: {added_rows}")

    # Ensure Sell Date is a true datetime64[ns] (Excel-friendly serial dates)
    if "Sell Date (EST)" in combined.columns:
        combined["Sell Date (EST)"] = pd.to_datetime(
            combined["Sell Date (EST)"], errors="coerce"
        )

    if combined.empty:
        print("Ledger is up to date; nothing new to add.")
        return

    # Recompute running total profit in chronological order
    combined = combined.sort_values(
        "Sell Date (EST)", ascending=True, kind="mergesort"
    ).reset_index(drop=True)
    if "Profit (g)" not in combined.columns:
        combined["Profit (g)"] = 0.0
    combined["Total Profit (g)"] = combined["Profit (g)"].cumsum()

    # Store newest first for convenience
    combined = combined.sort_values(
        "Sell Date (EST)", ascending=False, kind="mergesort"
    ).reset_index(drop=True)

    # Re-affirm datetime dtype prior to writing (avoid object dtype sneaking in)
    combined["Sell Date (EST)"] = pd.to_datetime(
        combined["Sell Date (EST)"], errors="coerce"
    )

    desired_order = [
        "Sell Date (EST)",
        "Item",
        "Quantity",
        "Avg Buy Price (g)",
        "Sell Price (g)",
        "Proceeds (g)",
        "Cost (g)",
        "Profit (g)",
        "ROI",
        "Hold Days",
        "Total Profit (g)",
        "sell_id",
        "item_id",
    ]
    existing_cols = [c for c in desired_order if c in combined.columns]
    extras = [c for c in combined.columns if c not in existing_cols]
    combined = combined[existing_cols + extras]

    # Build a daily summary for charting Total Profit over time
    summary = combined.copy()
    summary["Date"] = pd.to_datetime(summary["Sell Date (EST)"]).dt.floor("D")
    # Sort by true time, then take the last Total Profit per day (end-of-day equity)
    summary = summary.sort_values("Sell Date (EST)", ascending=True, kind="mergesort")
    summary_df = (
        summary.groupby("Date", as_index=False)
        .agg({"Total Profit (g)": "last"})
        .sort_values(by="Date", ascending=True)
        .reset_index(drop=True)
    )

    print(f"Writing {len(combined)} total rows to Excel...")

    # Write back (logical append; duplicates already removed) — single sheet only
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as writer:
        combined.to_excel(writer, index=False, sheet_name="Trades")
        # Apply a datetime number format to the Sell Date (EST) column for Excel x-axis friendliness
        ws = writer.sheets["Trades"]
        try:
            from openpyxl.utils import get_column_letter

            # Find the column index of "Sell Date (EST)"
            date_col_idx = (
                list(combined.columns).index("Sell Date (EST)") + 1
            )  # 1-based
            col_letter = get_column_letter(date_col_idx)
            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=date_col_idx,
                max_col=date_col_idx,
            ):
                for cell in row:
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
        except Exception:
            pass

    print(f"Updated ledger at: {xlsx_path} — added {len(new_df)} new rows.")


# ----------------------------- main -----------------------------


def get_profit():
    """
    Environment:
      - Set GW2_API_KEY in env (Tradingpost scope) or pass api_key below.
    Output:
      - Excel file 'gw2_trades.xlsx' in the current directory (append-only behavior).
    """
    print("Fetching transaction history...")
    gw2 = GW2()

    # 1) Fetch *history* (completed) buys & sells
    buys = fetch_all_transactions(gw2, scope="history", side="buys")
    buys = coalesce_buys_by_day(buys, tz="US/Eastern")
    sells = fetch_all_transactions(gw2, scope="history", side="sells")
    sells = coalesce_sells_by_day(sells, tz="US/Eastern")

    # Filter out items whose total coalesced buy value is < 10 gold
    threshold_copper = 10 * COPPER_PER_GOLD
    from collections import defaultdict as _dd

    buy_value_by_item = _dd(int)
    for b in buys:
        try:
            buy_value_by_item[int(b["item_id"])] += int(b["price"]) * int(b["quantity"])
        except Exception:
            pass
    allowed_items = {
        itm
        for itm, val in buy_value_by_item.items()
        if int(val) >= int(threshold_copper)
    }
    ignore_list = {
        19721,
        46736,
        46741,
        68063,
    }  # ecto, bolt of damask, spiritwood plank, amalgamated gemstone
    if allowed_items:
        # apply ignore list
        allowed_items = {itm for itm in allowed_items if itm not in ignore_list}
        before_buys, before_sells = len(buys), len(sells)
        buys = [b for b in buys if int(b["item_id"]) in allowed_items]
        sells = [s for s in sells if int(s["item_id"]) in allowed_items]
        print(
            f"Filtered by 10g min buy value: items kept={len(allowed_items)}; buys {before_buys}→{len(buys)}, sells {before_sells}→{len(sells)}"
        )
    else:
        print(
            "Warning: No items meet the 10g minimum coalesced buy value; skipping this filter."
        )

    print(f"Fetched {len(buys)} buy transactions and {len(sells)} sell transactions.")

    if not sells:
        print("No sell history returned by the API.")
        return

    # 2) Resolve item names
    item_ids = set(int(x["item_id"]) for x in buys) | set(
        int(x["item_id"]) for x in sells
    )
    name_lookup = fetch_item_names(gw2, item_ids) if item_ids else {}

    print(f"Resolved {len(name_lookup)} unique item names.")

    # 3) FIFO match & build rows (includes Hold Days)
    rows = fifo_match_and_rows(buys, sells, name_lookup)

    print(f"Built {len(rows)} matched sell records.")

    # 4) Append to Excel (no duplicates, recompute running total)
    out_path = os.path.abspath("gw2_trades.xlsx")
    append_to_excel(rows, out_path)


if __name__ == "__main__":
    get_profit()
